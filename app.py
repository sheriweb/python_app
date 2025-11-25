from fastapi import FastAPI, Response, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import re
import httpx
from bs4 import BeautifulSoup
import asyncio
import csv
import io
import os
from datetime import datetime

# Playwright for Maps
from playwright.async_api import async_playwright

app = FastAPI(title="Web Scraping Standalone API (Python)")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"\+?\d[\d\-()\s]{6,}\d")

# In-memory results store (no DB)
RESULTS: List[Dict[str, Any]] = []

# ---------- Models ----------
class UrlReq(BaseModel):
    url: str
    crawlDepth: Optional[int] = 1
    country: Optional[str] = None
    category: Optional[str] = None
    render: Optional[bool] = None  # if True, render with Playwright when needed

class SearchReq(BaseModel):
    prompt: Optional[str] = None
    category: Optional[str] = None
    country: Optional[str] = None
    limit: int = 10
    crawlDepth: Optional[int] = 1

class CollectReq(BaseModel):
    prompt: Optional[str] = None
    category: Optional[str] = None
    country: Optional[str] = None
    limit: int = 20

class MapsReq(BaseModel):
    mapsUrl: str
    limit: int = 50
    includeNoWebsite: bool = True
    aggressive: bool = False
    gridSweep: bool = False
    tiles: int = 0

# ---------- Helpers ----------
def normalize_url(u: str) -> str:
    try:
        u = (u or '').strip()
        if not u:
            return u
        from urllib.parse import urlparse
        parsed = urlparse(u)
        if not parsed.scheme:
            # default to https
            u = 'https://' + u
        return u
    except Exception:
        return u

async def fetch_text(url: str, timeout: int = 20) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }
    url = normalize_url(url)
    # Optional proxy support via PLAYWRIGHT_PROXY or HTTPS_PROXY/HTTP_PROXY
    proxies = None
    try:
        px = os.environ.get('PLAYWRIGHT_PROXY') or os.environ.get('HTTPS_PROXY') or os.environ.get('HTTP_PROXY')
        if px:
            proxies = {"all": px}
    except Exception:
        proxies = None
    # simple retry with backoff for transient errors
    for attempt in range(3):
        try:
            async with httpx.AsyncClient(follow_redirects=True, timeout=timeout, proxies=proxies) as client:
                r = await client.get(url, headers=headers)
                r.raise_for_status()
                return r.text
        except (httpx.ConnectError, httpx.ConnectTimeout, httpx.ReadTimeout):
            if attempt < 2:
                await asyncio.sleep(0.6 * (attempt + 1))
                continue
            raise
        except httpx.HTTPStatusError as e:
            # retry on 429/502/503/504
            if e.response is not None and e.response.status_code in (429, 502, 503, 504) and attempt < 2:
                await asyncio.sleep(0.8 * (attempt + 1))
                continue
            raise

async def fetch_rendered(url: str, timeout_ms: int = 15000) -> str:
    try:
        async with async_playwright() as p:
            proxy = os.environ.get("PLAYWRIGHT_PROXY")
            launch_kwargs = {"headless": True}
            if proxy:
                launch_kwargs["proxy"] = {"server": proxy}
            browser = await p.chromium.launch(**launch_kwargs)
            context = await browser.new_context(locale="en-US")
            page = await context.new_page()
            await page.goto(url, wait_until="domcontentloaded")
            await page.wait_for_timeout(600)
            # try to wait for network idle briefly
            try:
                await page.wait_for_load_state("networkidle", timeout=timeout_ms)
            except Exception:
                pass
            html = await page.content()
            await context.close(); await browser.close()
            return html
    except Exception:
        return ""

def deobfuscate_candidates(text: str) -> List[str]:
    # Normalize common patterns like [at], (at), {at}, dot, [.] etc.
    t = text.replace("\xa0", " ")
    patterns = [
        (re.compile(r"\s*(?:\[at\]|\(at\)|\{at\}|\sat\s|\sAT\s)\s*", re.I), "@"),
        (re.compile(r"\s*(?:\[dot\]|\(dot\)|\{dot\}|\sdot\s|\sDOT\s)\s*", re.I), "."),
        (re.compile(r"\s*\[\s*\.\s*\]\s*"), "."),
    ]
    for rx, repl in patterns:
        t = rx.sub(repl, t)
    # remove spaces around @ and .
    t = re.sub(r"\s*@\s*", "@", t)
    t = re.sub(r"\s*\.\s*", ".", t)
    # extract using standard regex now
    return list({e for e in EMAIL_RE.findall(t)})

def decode_cloudflare_cfemail(hexstr: str) -> Optional[str]:
    try:
        r = bytes.fromhex(hexstr)
        key = r[0]
        out = ''.join(chr(b ^ key) for b in r[1:])
        if '@' in out:
            return out
    except Exception:
        return None
    return None

def extract_jsonld_emails_phones(soup: BeautifulSoup) -> Dict[str, List[str]]:
    emails: List[str] = []
    phones: List[str] = []
    for tag in soup.find_all('script', type=lambda v: v and 'ld+json' in v):
        try:
            import json
            data = json.loads(tag.string or "{}")
            def walk(x):
                if isinstance(x, dict):
                    for k,v in x.items():
                        if k in ('email','emailAddress','contactPoint','contact','address','telephone','phone'):
                            if isinstance(v, str):
                                emails.extend(EMAIL_RE.findall(v))
                                phones.extend([m.group(0) for m in PHONE_RE.finditer(v)])
                            else:
                                walk(v)
                        else:
                            walk(v)
                elif isinstance(x, list):
                    for i in x: walk(i)
            walk(data)
        except Exception:
            continue
    return {"emails": list({*emails}), "phones": list({*phones})}

async def scrape_simple(url: str, crawl_depth: int = 1) -> Dict[str, Any]:
    html = await fetch_text(url)
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text(" ").replace("\xa0", " ")
    emails = list({e for e in EMAIL_RE.findall(text)})
    phones = list({m.group(0) for m in PHONE_RE.finditer(text)})
    title = (soup.title.string.strip() if soup.title and soup.title.string else None)
    desc_tag = soup.find("meta", attrs={"name": "description"})
    description = desc_tag.get("content").strip() if desc_tag and desc_tag.get("content") else None

    # pull from JSON-LD/microdata first
    try:
        agg = extract_jsonld_emails_phones(soup)
        if agg.get('emails'): emails = list({*emails, *agg['emails']})
        if agg.get('phones'): phones = list({*phones, *agg['phones']})
    except Exception:
        pass

    social = {"facebook": None, "instagram": None, "twitter": None, "linkedin": None, "other": []}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith("mailto:"):
            val = href[len("mailto:"):].strip()
            if val:
                emails.append(val)
            continue
        if href.startswith("tel:"):
            val = href[len("tel:"):].strip()
            if val:
                phones.append(val)
            continue
        if "facebook.com" in href and not social["facebook"]:
            social["facebook"] = href
        elif "instagram.com" in href and not social["instagram"]:
            social["instagram"] = href
        elif ("twitter.com" in href or "x.com" in href) and not social["twitter"]:
            social["twitter"] = href
        elif "linkedin.com" in href and not social["linkedin"]:
            social["linkedin"] = href
        elif len(social["other"]) < 10:
            social["other"].append(href)

    # Cloudflare-protected emails
    for el in soup.select('[data-cfemail]'):
        try:
            cf = el.get('data-cfemail')
            dec = decode_cloudflare_cfemail(cf)
            if dec: emails.append(dec)
        except Exception:
            pass

    # data-* attributes or obvious email text
    for el in soup.find_all(True):
        for attr in ('data-email','data-user','data-contact','data-mail','data_mail','data-contact-email'):
            v = el.get(attr)
            if v:
                emails.extend(EMAIL_RE.findall(str(v)))

    # deobfuscate from text blocks
    if not emails:
        emails = list({*emails, *deobfuscate_candidates(text)})

    if (not emails or not phones) and crawl_depth and crawl_depth > 1:
        try:
            from urllib.parse import urljoin
            for path in [
                "/contact", "/contact-us", "/contacts", "/get-in-touch",
                "/about", "/about-us", "/support", "/help"
            ]:
                try:
                    h2 = await fetch_text(urljoin(url, path))
                except Exception:
                    continue
                s2 = BeautifulSoup(h2, "lxml")
                t2 = s2.get_text(" ").replace("\xa0", " ")
                if not emails:
                    emails = list({*emails, *EMAIL_RE.findall(t2), *deobfuscate_candidates(t2)})
                if not phones:
                    phones = list({*phones, *[m.group(0) for m in PHONE_RE.finditer(t2)]})
                for a2 in s2.find_all("a", href=True):
                    hr = a2["href"]
                    if hr.startswith("mailto:"):
                        val = hr[len("mailto:"):].strip()
                        if val:
                            emails.append(val)
                    elif hr.startswith("tel:"):
                        val = hr[len("tel:"):].strip()
                        if val:
                            phones.append(val)
                if emails and phones:
                    break
        except Exception:
            pass

    if (not emails or not phones) and social.get("facebook"):
        try:
            fb_html = await fetch_text(social["facebook"])  
            fb_soup = BeautifulSoup(fb_html, "lxml")
            fb_text = fb_soup.get_text(" ")
            if not emails:
                emails = list({e for e in EMAIL_RE.findall(fb_text)})
            if not phones:
                phones = list({m.group(0) for m in PHONE_RE.finditer(fb_text)})
        except Exception:
            pass

    # Final fallback: render with Playwright once if nothing found
    if (not emails or not phones):
        rendered = await fetch_rendered(url)
        if rendered:
            rsoup = BeautifulSoup(rendered, "lxml")
            rtext = rsoup.get_text(" ")
            if not emails:
                emails = list({*emails, *EMAIL_RE.findall(rtext), *deobfuscate_candidates(rtext)})
            if not phones:
                phones = list({*phones, *[m.group(0) for m in PHONE_RE.finditer(rtext)]})

    return {
        "crawlDepth": crawl_depth or 1,
        "emails": emails,
        "phones": phones,
        "address": None,
        "socialLinks": social,
        "meta": {"title": title, "description": description},
    }

def push_result(doc: Dict[str, Any]):
    doc = {**doc}
    now = datetime.utcnow().isoformat()
    url = doc.get("url")
    # Upsert by URL: if exists, merge into existing instead of adding a new row
    if url:
        for i, existing in enumerate(RESULTS):
            if existing.get("url") == url:
                merged = {**existing, **doc}
                merged["createdAt"] = existing.get("createdAt", now)
                RESULTS[i] = merged
                return
    doc["createdAt"] = now
    RESULTS.insert(0, doc)

def make_query(prompt: Optional[str], category: Optional[str], country: Optional[str]) -> Optional[str]:
    return prompt or " ".join([x for x in [category, country] if x]) or None

async def search_collect(query: str, limit: int = 20) -> List[str]:
    urls: List[str] = []
    seen = set()

    def push(href: str):
        try:
            from urllib.parse import urlparse
            u = urlparse(href)
            if not u.scheme or not u.netloc:
                return
            host = u.netloc.replace("www.", "")
            origin = f"{u.scheme}://{u.netloc}"
            if host not in seen:
                seen.add(host)
                urls.append(origin)
        except Exception:
            pass

    q = httpx.QueryParams({"q": query})

    # DuckDuckGo pages
    for s in range(0, 150, 30):
        if len(urls) >= limit:
            break
        try:
            html = await fetch_text(f"https://html.duckduckgo.com/html/?{q}&s={s}")
            soup = BeautifulSoup(html, "lxml")
            for a in soup.select('a.result__a, .result__title a, a[href^="/l/"], a.result__url'):
                href = a.get('href')
                if not href:
                    continue
                from urllib.parse import urljoin, urlparse, parse_qs, unquote
                try:
                    u = urlparse(urljoin("https://duckduckgo.com", href))
                    uddg = parse_qs(u.query).get('uddg', [None])[0]
                    if uddg:
                        href = unquote(uddg)
                except Exception:
                    pass
                push(href)
        except Exception:
            pass

    # Bing fallback
    for first in range(1, 111, 10):
        if len(urls) >= limit:
            break
        try:
            html = await fetch_text(f"https://www.bing.com/search?q={query}&first={first}")
            soup = BeautifulSoup(html, "lxml")
            for a in soup.select('li.b_algo h2 a, h2 a'):
                href = a.get('href')
                if href:
                    push(href)
        except Exception:
            pass

    return urls[:limit]

async def maps_collect(maps_url: str, limit: int, include_no_site: bool, aggressive: bool, grid_sweep: bool, tiles: int):
    items = []
    seen = set()

    def key_of(it):
        if it.get("website"):
            return ("w", it["website"]) 
        return ("na", it.get("name"), it.get("address"))

    async with async_playwright() as p:
        proxy = os.environ.get("PLAYWRIGHT_PROXY")
        launch_kwargs = {"headless": True}
        if proxy:
            launch_kwargs["proxy"] = {"server": proxy}
        browser = await p.chromium.launch(**launch_kwargs)
        context = await browser.new_context(locale="en-US")
        page = await context.new_page()
        # Force English and US region to reduce consent/localization variance
        try:
            if ('hl=' not in maps_url) and ('gl=' not in maps_url):
                maps_url = maps_url + ("&" if "?" in maps_url else "?") + "hl=en&gl=us"
        except Exception:
            pass
        await page.goto(maps_url, wait_until="domcontentloaded")
        # Give Maps time to render the UI (especially on cold starts)
        await page.wait_for_timeout(1500)

        async def dismiss_consent():
            # Try common consent dialogs: EU/DE variants, different UIs
            selectors = [
                'button#L2AGLb',
                'form[action*="consent"] button',
                'button[aria-label^="Accept" i]',
                'button:has-text("I agree")',
                'button:has-text("Accept all")',
                'button:has-text("Ich stimme zu")',
                'button:has-text("Alle akzeptieren")',
            ]
            try:
                for sel in selectors:
                    try:
                        el = await page.query_selector(sel)
                        if el:
                            await el.click();
                            await page.wait_for_timeout(600)
                            break
                    except Exception:
                        continue
            except Exception:
                pass

        await dismiss_consent()

        async def collect_once():
            return await page.evaluate(r"""
(function () {
  function text(el) { return el ? (el.textContent || '').trim() : null; }
  function q(el, sel) { return el.querySelector(sel); }
  function qa(el, sel) { return Array.prototype.slice.call(el.querySelectorAll(sel)); }
  var out = [];
  var cards = document.querySelectorAll('.Nv2PK');
  for (var i = 0; i < cards.length; i++) {
    var c = cards[i];
    var n1 = q(c, '.qBF1Pd');
    var n2 = q(c, 'a.hfpxzc');
    var name = n1 ? text(n1) : (n2 ? n2.getAttribute('aria-label') : null);
    var website = null;
    var links = qa(c, 'a[role="link"], a[href]');
    for (var j = 0; j < links.length; j++) {
      var a = links[j];
      var txt = (a.textContent || '').toLowerCase();
      var href = a.getAttribute('href') || '';
      if (txt.indexOf('website') !== -1 || href.indexOf('/url?q=') === 0) {
        if (href.indexOf('/url?q=') === 0) {
          var qs = href.split('/url?q=')[1] || '';
          var target = qs.split('&')[0] || '';
          try { website = decodeURIComponent(target); } catch (e) { website = target; }
        } else if (href.indexOf('http') === 0) {
          website = href;
        }
        if (website) break;
      }
    }
    var blocks = qa(c, '.W4Efsd');
    var detailsArr = [];
    for (var k = 0; k < blocks.length; k++) { detailsArr.push(text(blocks[k]) || ''); }
    var details = detailsArr.join(' \n ');
    var pm = details.match(/\+?\d[\d\-()\s]{6,}\d/);
    var phone = pm ? pm[0] : null;
    var addrNode = q(c, '.W4Efsd span') || q(c, '.W4Efsd');
    var address = addrNode ? text(addrNode) : null;
    var rating = null, reviews = null;
    var rn = q(c, '[aria-label*="stars"], [aria-label*="rating"]');
    if (rn) {
      var al = rn.getAttribute('aria-label') || text(rn) || '';
      var mm = al.match(/([0-9]+(?:\.[0-9]+)?)/);
      if (mm) rating = parseFloat(mm[1]);
    }
    var rv = q(c, '.UY7F9') || q(c, 'span[aria-label*="reviews"]');
    if (rv) {
      var txt2 = rv.getAttribute('aria-label') || text(rv) || '';
      var mr = txt2.match(/([0-9][0-9,\.]*)/);
      if (mr) reviews = parseInt(mr[1].replace(/[^0-9]/g, ''));
    }
    out.push({ name: name, website: website, phone: phone, address: address, rating: rating, reviews: reviews });
  }
  return out;
})()
            """)

        list_sel = 'div[role="feed"], .m6QErb[aria-label]'

        async def collect_single_place():
            try:
                # Try to collect details from the place details panel when there is no feed (single place URL)
                return await page.evaluate(r"""
(function () {
  function text(el){ return el ? (el.textContent||'').trim() : null; }
  function q(sel){ return document.querySelector(sel); }
  function qa(sel){ return Array.prototype.slice.call(document.querySelectorAll(sel)); }
  var name = text(q('h1.DUwDvf span')) || text(q('h1.DUwDvf')) || text(q('[role="main"] h1')) || null;
  // Website button/link
  var website = null;
  var links = qa('a[role="link"], a[href]');
  for (var i=0;i<links.length;i++){
    var a = links[i];
    var t = (a.textContent||'').toLowerCase();
    var href = a.getAttribute('href')||'';
    if (t.indexOf('website') !== -1 || href.indexOf('/url?q=') === 0) {
      if (href.indexOf('/url?q=') === 0) {
        var qs = href.split('/url?q=')[1] || '';
        var target = qs.split('&')[0] || '';
        try { website = decodeURIComponent(target); } catch(e) { website = target; }
      } else if (href.indexOf('http') === 0) {
        website = href;
      }
      if (website) break;
    }
  }
  // Address & phone blocks often render inside the side panel details
  var panel = q('.m6QErb') || document.body;
  var blocks = qa('.Io6YTe, .W4Efsd, [data-item-id], [aria-label]', panel);
  var detailsTxt = blocks.map(function(n){ return text(n)||''; }).join('\n');
  var pm = detailsTxt.match(/\+?\d[\d\-()\s]{6,}\d/);
  var phone = pm ? pm[0] : null;
  var address = null;
  // Common address selectors
  var addrNode = q('[data-item-id="address"]') || q('button[data-item-id="address"]') || q('.W4Efsd span');
  if (!addrNode) {
    // fallback: first long-ish line that looks like an address
    var lines = detailsTxt.split(/\n+/).map(function(s){return s.trim();}).filter(Boolean);
    for (var j=0;j<lines.length;j++){
      if (lines[j].length > 15 && /[0-9]/.test(lines[j])) { address = lines[j]; break; }
    }
  } else {
    address = text(addrNode);
  }
  // rating and reviews
  var rating = null, reviews = null;
  var rn = q('[aria-label*="stars"], [aria-label*="rating"]');
  if (rn) {
    var al = rn.getAttribute('aria-label') || text(rn) || '';
    var mm = al.match(/([0-9]+(?:\.[0-9]+)?)/);
    if (mm) rating = parseFloat(mm[1]);
  }
  var rv = q('.UY7F9') || q('span[aria-label*="reviews"]');
  if (rv) {
    var txt2 = rv.getAttribute('aria-label') || text(rv) || '';
    var mr = txt2.match(/([0-9][0-9,\.]*)/);
    if (mr) reviews = parseInt(mr[1].replace(/[^0-9]/g, ''));
  }
  if (!name && !website && !phone && !address) return [];
  return [{ name: name, website: website, phone: phone, address: address, rating: rating, reviews: reviews }];
})()
                """)
            except Exception:
                return []

        async def scroll_once():
            try:
                feed = await page.query_selector(list_sel)
                if not feed:
                    await page.mouse.wheel(0, 1200)
                    return
                cards = await feed.query_selector_all('.Nv2PK')
                if cards:
                    await cards[-1].scroll_into_view_if_needed()
                await feed.evaluate("el => el.scrollBy(0, el.clientHeight - 60)")
            except Exception:
                pass

        async def click_search_area():
            try:
                el = await page.query_selector('[aria-label*="Search this area" i]')
                if el:
                    await el.click(); return
                for sel in ['button','a']:
                    nodes = await page.query_selector_all(sel)
                    for n in nodes:
                        txt = (await n.text_content() or '').strip().lower()
                        if 'search this area' in txt:
                            await n.click(); return
            except Exception:
                pass

        async def pan_zoom(step:int):
            try:
                await page.mouse.move(800 + step%3*5, 300 + step%2*5)
                await page.mouse.wheel(0, 220)
                if step % 4 == 0:
                    await page.keyboard.down('Control'); await page.mouse.wheel(0, -280); await page.keyboard.up('Control')
            except Exception:
                pass

        max_iters = 180 if aggressive else 90
        sweeps = max(0, min(tiles or 0, 50)) if grid_sweep else 0
        last_len = 0
        no_growth = 0

        # If there is no feed/list visible, try single-place extraction once
        try:
            feed_exists = await page.query_selector(list_sel)
        except Exception:
            feed_exists = None
        if not feed_exists:
            single = await collect_single_place()
            if single:
                for it in single:
                    k = key_of(it)
                    if k in seen: continue
                    if (not include_no_site) and (not it.get('website')): continue
                    seen.add(k); items.append(it)
                await context.close(); await browser.close()
                return items[:limit]

        for i in range(max_iters):
            chunk = await collect_once()
            for it in chunk:
                k = key_of(it)
                if k in seen: continue
                if (not include_no_site) and (not it.get('website')): continue
                seen.add(k); items.append(it)
            if len(items) >= limit: break
            await scroll_once()
            if aggressive:
                await click_search_area(); await pan_zoom(i)
            if grid_sweep and sweeps > 0 and (i % 6 == 3):
                try:
                    await page.mouse.down(); await page.mouse.move(420 + (i%3)*30, 220 + (i%2)*30, steps=10); await page.mouse.up();
                    await click_search_area(); sweeps -= 1
                except Exception:
                    pass
            await page.wait_for_timeout(700 + (i % 5) * 60)
            if len(items) == last_len:
                no_growth += 1
                if no_growth >= (10 if aggressive else 6): break
            else:
                no_growth = 0; last_len = len(items)

        await context.close(); await browser.close()
        return items[:limit]

# ---------- API Endpoints (no DB) ----------
@app.post("/api/scrape/url")
async def api_scrape_url(req: UrlReq):
    try:
        target = normalize_url(req.url)
        scraped = await scrape_simple(target, req.crawlDepth or 1)
        doc = {
            "url": target,
            "country": req.country,
            "category": req.category,
            **scraped,
        }
        push_result(doc)
        return {"success": True, "data": doc}
    except Exception as e:
        # Do not crash the server; report per-request failure
        return {"success": False, "message": f"Fetch failed: {e.__class__.__name__}"}

@app.post("/api/scrape/collect")
async def api_collect(req: CollectReq):
    q = make_query(req.prompt, req.category, req.country)
    if not q:
        return {"success": False, "message": "Provide prompt or category/country"}
    urls = await search_collect(q, max(1, min(req.limit, 500)))
    return {"success": True, "count": len(urls), "urls": urls}

@app.post("/api/scrape/search")
async def api_search(req: SearchReq):
    q = make_query(req.prompt, req.category, req.country)
    if not q:
        return {"success": False, "message": "Provide prompt or category/country"}
    urls = await search_collect(q, max(1, min(req.limit, 200)))
    results: List[Dict[str, Any]] = []
    sem = asyncio.Semaphore(5)

    async def worker(u: str):
        async with sem:
            try:
                s = await scrape_simple(u, req.crawlDepth or 1)
                doc = {"url": u, "country": req.country, "category": req.category, **s}
                results.append(doc)
            except Exception:
                pass

    await asyncio.gather(*(worker(u) for u in urls))
    for d in results: push_result(d)
    return {"success": True, "count": len(results), "data": results}

@app.post("/api/scrape/maps")
async def api_maps(req: MapsReq):
    data = await maps_collect(req.mapsUrl, max(1, min(req.limit, 500)), req.includeNoWebsite, req.aggressive, req.gridSweep, req.tiles)
    # Do not push into RESULTS automatically; only when user clicks Get details per site (client)
    return {"success": True, "count": len(data), "data": data}

@app.post("/api/scrape/import-csv")
async def api_import_csv(file: UploadFile = File(...), replace: bool = False):
    raw = await file.read()
    filename = (file.filename or '').lower()

    # Optionally clear previous results
    if replace:
        RESULTS.clear()

    def normalize_headers(row: Dict[str, Any]) -> Dict[str, Any]:
        return { (k or '').strip().lower(): (v if v is not None else '') for k, v in row.items() }

    def pick(row: Dict[str, Any], keys) -> str:
        for k in keys:
            if k in row and str(row[k]).strip():
                return str(row[k]).strip()
        # fuzzy: try any header containing the token
        for k in row.keys():
            lk = k.lower()
            for token in keys:
                if token in lk and str(row[k]).strip():
                    return str(row[k]).strip()
        return ''

    def to_row_docs(rows: List[Dict[str, Any]]):
        cnt = 0
        for r in rows:
            row = normalize_headers(r)
            website = pick(row, ["website","url","site","domain","homepage"]) 
            if not website:
                continue
            name = pick(row, ["name","title","business name","company"]) or None
            addr1 = pick(row, ["address_1","address1","address line 1","address"]) or ''
            addr2 = pick(row, ["address_2","address2","address line 2","city","state","postcode","zip"]) or ''
            address = ' '.join([x for x in [addr1, addr2] if x]) or None
            phone_raw = pick(row, ["phone_number","phone","mobile","tel"]) or ''
            rating_raw = pick(row, ["rating","stars"]) or ''
            reviews_raw = pick(row, ["reviews_count","reviews","review count"]) or ''
            rating = None
            if rating_raw:
                try: rating = float(str(rating_raw).replace(',', '.'))
                except Exception: rating = None
            reviews = None
            if reviews_raw:
                try: reviews = int(re.sub(r"[^0-9]", "", str(reviews_raw)))
                except Exception: reviews = None
            doc = {
                "url": website,
                "meta": {"title": name, "description": None},
                "emails": [],
                "phones": [phone_raw] if phone_raw else [],
                "address": address,
                "rating": rating,
                "reviews": reviews,
                "socialLinks": {},
            }
            push_result(doc)
            cnt += 1
        return cnt

    # If content looks like XLSX (ZIP 'PK' magic), treat as XLSX even if extension missing
    looks_xlsx = raw[:2] == b'PK'
    # CSV path
    if not filename.endswith('.xlsx') and not looks_xlsx:
        try:
            text = raw.decode('utf-8-sig', errors='ignore')
        except Exception:
            text = raw.decode('latin-1', errors='ignore')
        f = io.StringIO(text)
        reader = csv.DictReader(f)
        rows = [dict(r) for r in reader]
        count = to_row_docs(rows)
        return {"success": True, "count": count}

    # XLSX path
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active

        expected = ["website","url","site","domain","homepage","name","address","phone","rating","reviews"]
        header_row_idx = None
        best_score = -1
        # scan first 20 rows to find the most likely header row
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20)):
            vals = [str(c.value or '').strip().lower() for c in row]
            score = sum(any(tok in v for tok in expected) for v in vals)
            if score > best_score:
                best_score = score
                header_row_idx = idx + 1
        if not header_row_idx:
            header_row_idx = 1

        headers = [str(c.value or '').strip() for c in ws[header_row_idx]]
        data_rows = []
        for row_cells in ws.iter_rows(min_row=header_row_idx+1):
            # stop if row mostly empty
            if all((c.value is None or str(c.value).strip()=="") and (c.hyperlink is None or not c.hyperlink.target) for c in row_cells):
                continue
            row_map = {}
            for i, cell in enumerate(row_cells):
                key = headers[i] if i < len(headers) else f"col{i}"
                val = cell.value if cell.value is not None else ''
                if cell.hyperlink is not None and cell.hyperlink.target:
                    lk = key.lower()
                    if any(tok in lk for tok in ["website","url","site","domain","homepage"]):
                        val = cell.hyperlink.target
                row_map[key] = val
            data_rows.append(row_map)

        count = to_row_docs(data_rows)
        sample_web = []
        for r in data_rows:
            nr = { (k or '').strip().lower(): (v if v is not None else '') for k,v in r.items() }
            w = None
            for k in ["website","url","site","domain","homepage"]:
                if k in nr and str(nr[k]).strip():
                    w = str(nr[k]).strip(); break
            if not w:
                for k,v in nr.items():
                    if any(tok in k for tok in ["website","url","site","domain","homepage"]) and str(v).strip():
                        w = str(v).strip(); break
            if w:
                sample_web.append(w)
            if len(sample_web) >= 3:
                break
        return {"success": True, "count": count, "headers": headers, "headerRow": header_row_idx, "sampleWebsites": sample_web}
    except Exception as e:
        return {"success": False, "message": f"Failed to read xlsx: {e}"}

@app.get("/api/scrape/results")
async def api_results(page: int = 1, limit: int = 10, format: str = "json"):
    fmt = (format or "").lower()
    total = len(RESULTS)
    if fmt == "csv":
        # For CSV export, do not cap limit; allow exporting all when a very large limit is requested.
        if limit >= 100000 or page <= 0 or limit >= total:
            rows = RESULTS
        else:
            start = max(0, (max(1, page)-1) * max(1, limit))
            end = start + max(1, limit)
            rows = RESULTS[start:end]
        headers = ["url","title","description","emails","phones","facebook","instagram","twitter","linkedin","country","category","createdAt"]
        sio = io.StringIO(); w = csv.writer(sio)
        w.writerow(headers)
        for r in rows:
            title = r.get("meta",{}).get("title") or ""
            desc = r.get("meta",{}).get("description") or ""
            emails = " ".join(r.get("emails") or [])
            phones = " ".join(r.get("phones") or [])
            sl = r.get("socialLinks", {})
            w.writerow([
                r.get("url",""), title, desc, emails, phones,
                sl.get("facebook",""), sl.get("instagram",""), sl.get("twitter",""), sl.get("linkedin",""),
                r.get("country",""), r.get("category",""), r.get("createdAt",""),
            ])
        return Response(content=sio.getvalue(), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=results.csv"})
    # JSON path: cap for UI pagination (allow up to 1000)
    page = max(1, page); limit = max(1, min(1000, limit))
    start = (page-1) * limit; end = start + limit
    rows = RESULTS[start:end]
    return {"success": True, "total": total, "page": page, "limit": limit, "data": rows}

@app.post("/api/scrape/clear")
async def api_clear():
    RESULTS.clear()
    return {"success": True, "cleared": True, "total": len(RESULTS)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8010)
